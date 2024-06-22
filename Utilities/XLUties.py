import openpyxl


def ReadData(file, sheet_name, ro, col):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row=ro, column=col).value


def WriteData(file, sheet_name, ro, col, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row=ro, column=col).value = data
    workbook.save(file)


def RowCount(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row()
